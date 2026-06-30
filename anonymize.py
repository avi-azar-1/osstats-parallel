import sys
import json
import os
from openpyxl import load_workbook


def anonymize(input_path):
    if not os.path.isfile(input_path):
        print(f"Error: file not found: {input_path}")
        sys.exit(1)

    wb = load_workbook(input_path)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    try:
        col_idx = header_row.index("NodeId") + 1
    except ValueError:
        print("Error: 'NodeId' column not found in header row")
        sys.exit(1)

    mapping = {}
    counter = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val is not None and val not in mapping:
            counter += 1
            mapping[val] = f"Node-{counter}"

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value in mapping:
            cell.value = mapping[cell.value]

    stem = input_path.rsplit(".", 1)[0]
    output_path = f"{stem}_anonymized.xlsx"
    mapping_path = f"{stem}_mapping.json"

    wb.save(output_path)
    with open(mapping_path, "w") as f:
        json.dump(mapping, f, indent=2)

    print(f"Anonymized {counter} unique NodeId values across {ws.max_row - 1} rows")
    print(f"Output: {output_path}")
    print(f"Mapping: {mapping_path}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python anonymize.py <input.xlsx>")
        sys.exit(1)
    anonymize(sys.argv[1])
